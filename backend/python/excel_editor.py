#!/usr/bin/env python3
"""
Excel編集スクリプト（openpyxl使用）

テンプレートExcelにPDF解析データを転記し、新しいExcelファイルを生成します。
数式は保持され、自動計算されます。

使用法:
    python3 excel_editor.py <company_name> <template_path> <output_path> <data_json>

引数:
    company_name: 取引先名（ネクストビッツ or オフ・ビート・ワークス）
    template_path: テンプレートExcelファイルのパス
    output_path: 出力Excelファイルのパス
    data_json: PDF解析データ（JSON文字列）

出力:
    編集済みExcelファイルのパス（標準出力）

処理ルール（docs/processing_rules.md）:
    ネクストビッツ:
        注文書シート:
            - AC2: 発行日（見積書ファイル名から処理対象月を抽出し、当月1日を入力）
            - R18: 数量（見積書から、「1式」→「1」）
            - T18: 単価（見積書から）
        検収書シート:
            - R20: 数量（注文書シートR18と同じ）
            - T20: 単価（注文書シートT18と同じ）

    オフ・ビート・ワークス:
        注文書シート:
            - AC2: 発行日（注文請書の発行日）
            - AA17: 摘要（見積書番号を「見積番号：NNNNNNN」形式で入力）
            - C18~: 件名（請求書の品目、先頭に「・」を付ける）
            - R18~: 数量（請求書から、複数行対応）
            - T18~: 単価（請求書から）
        検収書シート:
            - C20~: 件名（注文書シートC18~と同じ）
            - R20~: 数量（注文書シートR18~と同じ）
            - T20~: 単価（注文書シートT18~と同じ）
"""

import sys
import json
import openpyxl
from datetime import datetime
from dateutil.relativedelta import relativedelta
from typing import Dict, Any, List


def edit_nextbits_excel(wb: openpyxl.Workbook, data: Dict[str, Any]) -> datetime:
    """
    ネクストビッツ様のExcel編集

    実際のExcel構造（テラ【株式会社ネクストビッツ御中】注文検収書_YYMM.xlsx）:
    - シート: 注文書, 検収書

    注文書シート:
    - AC2: 発行日（編集: 見積書ファイル名から処理対象月を抽出し、当月1日を入力）
    - AC3: 注文番号（数式: 自動計算）
    - R18: 数量（編集: 見積書から）
    - T18: 単価（編集: 見積書から）
    - W18: 金額（数式: 自動計算）
    - AA17: 摘要（数式: 自動計算）

    検収書シート:
    - AC4: 検収番号（数式: 自動計算）
    - AC5: 検収日（数式: 自動計算）
    - R20: 数量（編集: 見積書から）
    - T20: 単価（編集: 見積書から）
    - W20: 金額（数式: 自動計算）
    - AA19: 摘要（数式: 自動計算）

    Returns:
        発行日（TEXT関数のキャッシュ値計算に使用）
    """
    ws_order = wb["注文書"]
    ws_inspection = wb["検収書"]

    # 見積書データを取得
    estimate_data = data.get('estimate', {})
    quantity = estimate_data.get('quantity', 1)
    unit_price = estimate_data.get('unit_price', 0)

    # === 注文書シート編集 ===

    # 発行日: 見積書ファイル名から処理対象月を抽出（TRR-YY-MMM → 20YY年MM月1日）
    # estimate_filenameはprocessServiceから渡される
    estimate_filename = data.get('estimate_filename', '')
    issue_date = None

    if estimate_filename:
        # ファイル名パターン: TRR-YY-MMM_お見積書.pdf（例: TRR-25-007 → 2025年7月1日）
        import re
        match = re.search(r'TRR-(\d{2})-(\d{3})', estimate_filename)
        if match:
            year = 2000 + int(match.group(1))  # YY → 20YY
            month = int(match.group(2))  # MMM → MM（007 → 7）
            issue_date = datetime(year, month, 1)

    if not issue_date:
        # ファイル名から取得できない場合は今月の1日
        issue_date = datetime.now().replace(day=1)

    # 発行日を更新
    ws_order['AC2'] = issue_date

    # 数量を更新（R18）
    ws_order['R18'] = quantity

    # 単価を更新（T18）
    ws_order['T18'] = unit_price

    # === 件名チェック（C18, C20） ===
    # 処理ルール: 見積書の件名が「yyyy年mm月作業：Telemasシステム改修作業等」の場合、
    # C18が「　Telemas作業(システム改修等)」になっているかチェック
    subject = estimate_data.get('subject', '')
    expected_c18 = '　Telemas作業(システム改修等)'  # 先頭に全角スペース

    if subject:
        # 見積書の件名パターンをチェック
        import re
        telemas_pattern = re.match(r'\d{4}年\d{2}月作業[：:]Telemasシステム改修作業等', subject)

        if telemas_pattern:
            # Telemasパターンの場合、C18が正しいかチェック
            current_c18 = ws_order['C18'].value
            if current_c18 != expected_c18:
                raise ValueError(
                    f"件名不一致エラー: C18の値が正しくありません。\n"
                    f"期待値: 「{expected_c18}」\n"
                    f"実際値: 「{current_c18}」"
                )
            # C20も同様にチェック
            current_c20 = ws_inspection['C20'].value
            if current_c20 != expected_c18:
                raise ValueError(
                    f"件名不一致エラー: C20の値が正しくありません。\n"
                    f"期待値: 「{expected_c18}」\n"
                    f"実際値: 「{current_c20}」"
                )
        else:
            # 未対応の件名パターン
            raise ValueError(f"未対応の件名パターンです: 「{subject}」")

    # === 検収書シート編集 ===
    # 注: 検収書のAC4, AC5, AA19は注文書シートを参照する数式なので自動更新される

    # 数量を更新（R20）
    ws_inspection['R20'] = quantity

    # 単価を更新（T20）
    ws_inspection['T20'] = unit_price

    # 発行日を返す（TEXT関数のキャッシュ値計算に使用）
    return issue_date


def edit_offbeat_excel(wb: openpyxl.Workbook, data: Dict[str, Any]) -> datetime:
    """
    オフ・ビート・ワークス様のExcel編集

    実際のExcel構造（テラ【株式会社オフ・ビート・ワークス御中】注文検収書_YYMM.xlsx）:
    - シート: 注文書, 検収書

    注文書シート:
    - AC2: 発行日（編集: 注文請書の発行日）
    - AC3: 注文番号（数式: 自動計算）
    - AA17: 摘要（編集: 見積書番号を「見積番号：NNNNNNN」形式で入力）
    - C18~: 件名（編集: 請求書の品目、先頭に「・」を付ける）
    - R18~: 数量（編集: 請求書の明細から、複数行対応）
    - T18~: 単価（編集: 請求書の明細から）

    検収書シート:
    - AC4: 検収番号（数式: 自動計算）
    - AC5: 検収日（数式: 自動計算）
    - AA19: 摘要（数式: =注文書!AA17 で自動計算）
    - C20~: 件名（編集: 請求書の品目、先頭に「・」を付ける）
    - R20~: 数量（編集: 請求書の明細から）
    - T20~: 単価（編集: 請求書の明細から）

    Returns:
        発行日（TEXT関数のキャッシュ値計算に使用）
    """
    ws_order = wb["注文書"]
    ws_inspection = wb["検収書"]

    # 各PDFからのデータを取得
    estimate_data = data.get('estimate', {})
    invoice_data = data.get('invoice', {})
    order_confirmation_data = data.get('order_confirmation', {})

    # === 注文書シート編集 ===

    # 発行日（注文請書の発行日から）
    issue_date = None
    issue_date_str = order_confirmation_data.get('issue_date', '')
    if issue_date_str:
        try:
            issue_date = datetime.strptime(issue_date_str, '%Y-%m-%d')
            ws_order['AC2'] = issue_date
        except ValueError:
            pass  # パースエラーの場合は元の値を保持
    if issue_date is None:
        # 発行日が取得できない場合は今月の1日
        issue_date = datetime.now().replace(day=1)

    # 見積書番号（摘要用）: 「見積番号：NNNNNNN」形式で入力
    estimate_number = estimate_data.get('estimate_number', '')
    if estimate_number:
        ws_order['AA17'] = f"見積番号：{estimate_number}"

    # 明細行（請求書から）
    items: List[Dict[str, Any]] = invoice_data.get('items', [])

    # まず前月の明細データをクリア（最大20行分、行18〜37をクリア）
    MAX_DETAIL_ROWS = 20
    for i in range(MAX_DETAIL_ROWS):
        row_order = 18 + i
        row_inspection = 20 + i
        # 注文書シートのクリア（C, R, T, W列）
        ws_order.cell(row=row_order, column=3).value = None   # C列（件名）
        ws_order.cell(row=row_order, column=18).value = None  # R列（数量）
        ws_order.cell(row=row_order, column=20).value = None  # T列（単価）
        ws_order.cell(row=row_order, column=23).value = None  # W列（金額）
        # 検収書シートのクリア（C, R, T, W列）
        ws_inspection.cell(row=row_inspection, column=3).value = None   # C列（件名）
        ws_inspection.cell(row=row_inspection, column=18).value = None  # R列（数量）
        ws_inspection.cell(row=row_inspection, column=20).value = None  # T列（単価）
        ws_inspection.cell(row=row_inspection, column=23).value = None  # W列（金額）

    if items:
        # 明細行を編集（最大行数は適宜調整）
        for i, item in enumerate(items):
            row_order = 18 + i  # 注文書の明細開始行（R18, C18から）
            row_inspection = 20 + i  # 検収書の明細開始行（R20, C20から）

            # 注文書シート
            ws_order.cell(row=row_order, column=18).value = item.get('quantity', 1)  # R列（18列目）
            ws_order.cell(row=row_order, column=20).value = item.get('unit_price', 0)  # T列（20列目）
            # W列に金額計算の数式を追加（=T[行]*R[行]）
            ws_order.cell(row=row_order, column=23).value = f"=T{row_order}*R{row_order}"  # W列（23列目）

            # 件名（品目に「・」を付ける）
            item_name = item.get('name', '')
            if item_name and not item_name.startswith('・'):
                item_name = '・' + item_name
            ws_order.cell(row=row_order, column=3).value = item_name  # C列（3列目）

            # 検収書シート
            ws_inspection.cell(row=row_inspection, column=18).value = item.get('quantity', 1)  # R列
            ws_inspection.cell(row=row_inspection, column=20).value = item.get('unit_price', 0)  # T列
            # W列に金額計算の数式を追加（=T[行]*R[行]）
            ws_inspection.cell(row=row_inspection, column=23).value = f"=T{row_inspection}*R{row_inspection}"  # W列（23列目）

            # 件名（品目に「・」を付ける）
            ws_inspection.cell(row=row_inspection, column=3).value = item_name  # C列

        # 明細最終行の次の行に「以下、余白」を入力
        last_row_order = 18 + len(items)
        last_row_inspection = 20 + len(items)
        ws_order.cell(row=last_row_order, column=3).value = "以下、余白"
        ws_inspection.cell(row=last_row_inspection, column=3).value = "以下、余白"

    else:
        # 明細がない場合（単一明細の場合）
        # 請求書の合計から単一明細として処理
        subtotal = invoice_data.get('subtotal', 0)
        if subtotal > 0:
            ws_order['R18'] = 1
            ws_order['T18'] = subtotal
            ws_order['W18'] = "=T18*R18"  # W列に金額計算の数式を追加

            ws_inspection['R20'] = 1
            ws_inspection['T20'] = subtotal
            ws_inspection['W20'] = "=T20*R20"  # W列に金額計算の数式を追加

            # 「以下、余白」を入力
            ws_order['C19'] = "以下、余白"
            ws_inspection['C21'] = "以下、余白"

    # 発行日を返す（TEXT関数のキャッシュ値計算に使用）
    return issue_date


def validate_totals(wb: openpyxl.Workbook, data: Dict[str, Any], company_name: str) -> Dict[str, Any]:
    """
    Excel計算結果と請求書の金額を照合

    Returns:
        検証結果（一致/不一致、差分情報）
    """
    invoice_data = data.get('invoice', {})
    invoice_total = invoice_data.get('total', 0)
    invoice_subtotal = invoice_data.get('subtotal', 0)
    invoice_tax = invoice_data.get('tax', 0)

    # Excel側の合計を取得（openpyxlは数式の計算結果を取得できないため、
    # 実際の検証はLibreOfficeでPDF生成時に行う）

    return {
        "invoice_total": invoice_total,
        "invoice_subtotal": invoice_subtotal,
        "invoice_tax": invoice_tax,
        "note": "数式の計算結果はLibreOfficeでPDF生成時に検証"
    }


def restore_drawing_from_template(template_path: str, output_path: str, issue_date: datetime = None, company_name: str = None) -> None:
    """
    テンプレートからopenpyxlが削除/変更したファイルを復元する

    openpyxlは以下のファイルを適切に処理できない:
    - [Content_Types].xml: ContentType定義が変更される
    - _rels/.rels: ルートリレーションが変更される
    - xl/_rels/workbook.xml.rels: ワークブックリレーションが変更される
    - xl/workbook.xml: ワークブック定義が変更される
    - xl/calcChain.xml: 計算チェーンが削除される
    - xl/sharedStrings.xml: 共有文字列が削除される
    - xl/drawings/: Office 2014以降の拡張情報が削除される
    - xl/printerSettings/: プリンタ設定が削除される
    - xl/worksheets/_rels/: シートのリレーションファイルが変更される

    また、openpyxlは数式のキャッシュ値（計算結果）を保持しないため、
    TEXT関数などのキャッシュ値をPythonで計算して設定する。

    この関数は、テンプレートからこれらのファイルをそのまま出力ファイルにコピーし、
    シートXML（データ）のみを処理後ファイルから取得することで、Excelでの修復エラーを防ぐ。

    Args:
        template_path: テンプレートExcelファイルのパス
        output_path: 出力Excelファイルのパス（修復対象）
        issue_date: 発行日（TEXT関数のキャッシュ値計算用）
    """
    import zipfile
    import tempfile
    import shutil
    import os
    import re

    # テンプレートから復元するファイル
    # openpyxlが壊す/削除するファイルのみをテンプレートから取得
    # 注意: calcChain.xmlは復元しない（openpyxlの編集内容と不整合になり、数式が再計算されなくなる）
    files_from_template = [
        # drawingsフォルダ内のファイル（Office拡張情報）
        'xl/drawings/drawing1.xml',
        'xl/drawings/_rels/drawing1.xml.rels',
        # printerSettingsフォルダ（削除される）
        'xl/printerSettings/printerSettings1.bin',
        'xl/printerSettings/printerSettings2.bin',
        # シートのリレーションファイル（rIdが変更される）
        'xl/worksheets/_rels/sheet1.xml.rels',
        'xl/worksheets/_rels/sheet2.xml.rels',
        # sharedStrings（削除される場合がある）
        'xl/sharedStrings.xml',
        # Content_Typesとrels（変更される）
        '[Content_Types].xml',
        '_rels/.rels',
        'xl/_rels/workbook.xml.rels',
        # workbook.xml（定義が変更される）
        'xl/workbook.xml',
        # docProps（メタデータが変更される）
        'docProps/app.xml',
        'docProps/core.xml',
    ]

    # 削除するファイル（復元もしない、存在させない）
    # calcChain.xmlを削除することで、Excelの修復ダイアログを回避
    # openpyxlが作成するcalcChain.xmlは行数が少ない場合に破損するため、
    # 削除しておき、Excelが開く際に自動再生成させる
    files_to_remove = [
        'xl/calcChain.xml',
    ]

    template_files = {}
    processed_files = {}
    template_drawing_rids = {}

    try:
        # テンプレートファイルを読み込み
        with zipfile.ZipFile(template_path, 'r') as template_zip:
            for name in template_zip.namelist():
                template_files[name] = template_zip.read(name)

            # テンプレートのシートXMLからdrawing参照のrIdを取得
            for name in template_zip.namelist():
                if name.startswith('xl/worksheets/sheet') and name.endswith('.xml'):
                    content = template_files[name].decode('utf-8')
                    match = re.search(r'<drawing[^>]*r:id="(rId\d+)"', content)
                    if match:
                        template_drawing_rids[name] = match.group(1)

        # 処理後ファイルを読み込み
        with zipfile.ZipFile(output_path, 'r') as output_zip:
            for name in output_zip.namelist():
                processed_files[name] = output_zip.read(name)

    except Exception as e:
        print(f"[restore_drawing] Error reading files: {e}", file=sys.stderr)
        raise

    # 出力ファイルを更新
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_path = tmp_file.name

        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as new_zip:
            # 処理後ファイルのすべてのファイルをベースにする
            for name, content in processed_files.items():
                # 削除対象のファイルはスキップ
                if name in files_to_remove:
                    continue

                # テンプレートから復元すべきファイルの場合
                if name in files_from_template and name in template_files:
                    content = template_files[name]

                # シートXMLの場合、drawing参照のrIdを修正（テンプレートの正しいrIdに合わせる）
                if name in template_drawing_rids:
                    content_str = content.decode('utf-8')
                    correct_rid = template_drawing_rids[name]
                    # drawing参照のrIdを正しい値に置換
                    content_str = re.sub(
                        r'(<drawing[^>]*r:id=")rId\d+(")',
                        rf'\g<1>{correct_rid}\2',
                        content_str
                    )
                    content = content_str.encode('utf-8')

                # シートXML（sheet1.xml = 注文書）の場合、TEXT関数のキャッシュ値を設定
                # openpyxlは数式のキャッシュ値を保持しないため、Pythonで計算して設定する
                if name == 'xl/worksheets/sheet1.xml' and issue_date is not None:
                    content_str = content.decode('utf-8')

                    # AC3: 注文番号（取引先ごとに異なる形式）
                    # ネクストビッツ: =TEXT(AC2,"yyyymmdd")&"-01" → "20250801-01"
                    # オフビートワークス: =TEXT(AC2,"yyyymmdd")&"-02" → "20250801-02"
                    order_suffix = '-02' if company_name == 'オフ・ビート・ワークス' else '-01'
                    ac3_value = issue_date.strftime('%Y%m%d') + order_suffix
                    # <v></v> または <v/> を <v>計算結果</v> に置換
                    # AC3セルのパターン: <c r="AC3" ...><f>...</f><v></v></c>
                    content_str = re.sub(
                        r'(<c r="AC3"[^>]*>)(<f>[^<]*</f>)<v></v>(</c>)',
                        rf'\1\2<v>{ac3_value}</v>\3',
                        content_str
                    )
                    # <v/>の場合も対応
                    content_str = re.sub(
                        r'(<c r="AC3"[^>]*>)(<f>[^<]*</f>)<v/>(</c>)',
                        rf'\1\2<v>{ac3_value}</v>\3',
                        content_str
                    )
                    # t="str"属性を追加（数式の結果が文字列であることを示す）
                    # <c r="AC3" s="47"> → <c r="AC3" s="47" t="str">
                    content_str = re.sub(
                        r'(<c r="AC3" s="\d+")(?! t="str")([^>]*>)',
                        r'\1 t="str"\2',
                        content_str
                    )

                    # C17: 明細タイトル（取引先ごとに異なる形式）
                    # ネクストビッツ: =TEXT(AC2,"yyyy年mm月分作業費") → "2025年08月分作業費"
                    # オフビートワークス: =TEXT(AC2,"yyyy年mm月作業費") → "2025年08月作業費"（「分」なし）
                    if company_name == 'オフ・ビート・ワークス':
                        c17_value = issue_date.strftime('%Y年%m月作業費')
                    else:
                        c17_value = issue_date.strftime('%Y年%m月分作業費')
                    content_str = re.sub(
                        r'(<c r="C17"[^>]*>)(<f>[^<]*</f>)<v></v>(</c>)',
                        rf'\1\2<v>{c17_value}</v>\3',
                        content_str
                    )
                    content_str = re.sub(
                        r'(<c r="C17"[^>]*>)(<f>[^<]*</f>)<v/>(</c>)',
                        rf'\1\2<v>{c17_value}</v>\3',
                        content_str
                    )
                    # t="str"属性を追加
                    content_str = re.sub(
                        r'(<c r="C17" s="\d+")(?! t="str")([^>]*>)',
                        r'\1 t="str"\2',
                        content_str
                    )

                    content = content_str.encode('utf-8')

                # シートXML（sheet2.xml = 検収書）の場合も同様にキャッシュ値を設定
                if name == 'xl/worksheets/sheet2.xml' and issue_date is not None:
                    content_str = content.decode('utf-8')

                    # C19: 明細タイトル（取引先ごとに異なる形式）
                    # ネクストビッツ: =TEXT(注文書!$AC$2,"yyyy年mm月分作業費") → "2025年08月分作業費"
                    # オフビートワークス: =TEXT(注文書!$AC$2,"yyyy年mm月作業費") → "2025年08月作業費"（「分」なし）
                    if company_name == 'オフ・ビート・ワークス':
                        c19_value = issue_date.strftime('%Y年%m月作業費')
                    else:
                        c19_value = issue_date.strftime('%Y年%m月分作業費')
                    content_str = re.sub(
                        r'(<c r="C19"[^>]*>)(<f>[^<]*</f>)<v></v>(</c>)',
                        rf'\1\2<v>{c19_value}</v>\3',
                        content_str
                    )
                    content_str = re.sub(
                        r'(<c r="C19"[^>]*>)(<f>[^<]*</f>)<v/>(</c>)',
                        rf'\1\2<v>{c19_value}</v>\3',
                        content_str
                    )
                    # t="str"属性を追加
                    content_str = re.sub(
                        r'(<c r="C19" s="\d+")(?! t="str")([^>]*>)',
                        r'\1 t="str"\2',
                        content_str
                    )

                    content = content_str.encode('utf-8')

                # workbook.xmlの場合、強制再計算フラグを追加
                # calcPr要素にfullCalcOnLoad="1"を設定してExcelがファイルを開いた時に再計算させる
                if name == 'xl/workbook.xml':
                    content_str = content.decode('utf-8')
                    # calcPr要素が存在する場合、fullCalcOnLoadを追加
                    if '<calcPr' in content_str:
                        # fullCalcOnLoad属性がない場合のみ追加
                        if 'fullCalcOnLoad' not in content_str:
                            content_str = re.sub(
                                r'(<calcPr)',
                                r'\1 fullCalcOnLoad="1"',
                                content_str
                            )
                    else:
                        # calcPr要素がない場合、workbook終了タグの前に追加
                        content_str = re.sub(
                            r'(</workbook>)',
                            r'<calcPr fullCalcOnLoad="1"/>\1',
                            content_str
                        )
                    content = content_str.encode('utf-8')

                new_zip.writestr(name, content)

            # テンプレートにあって処理後ファイルにないファイルを追加
            # （openpyxlが削除したファイルを復元）
            # ただし削除対象のファイルは復元しない
            for name, content in template_files.items():
                if name not in processed_files and name not in files_to_remove:
                    new_zip.writestr(name, content)

        # 元のファイルを置き換え
        shutil.move(tmp_path, output_path)

    except Exception as e:
        if tmp_path and os.path.exists(tmp_path):
            os.remove(tmp_path)
        print(f"[restore_drawing] Error restoring files: {e}", file=sys.stderr)
        raise


def edit_excel(company_name: str, template_path: str, output_path: str, data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Excelテンプレートにデータを転記

    Args:
        company_name: 取引先名
        template_path: テンプレートExcelファイルのパス
        output_path: 出力Excelファイルのパス
        data: PDF解析データ（辞書型）

    Returns:
        編集結果（パス、検証結果を含む）
    """
    try:
        # テンプレートExcelを読み込み
        wb = openpyxl.load_workbook(template_path)

        # 取引先ごとの編集処理（発行日を返す）
        issue_date = None
        if company_name == "ネクストビッツ":
            issue_date = edit_nextbits_excel(wb, data)
        elif company_name == "オフ・ビート・ワークス":
            issue_date = edit_offbeat_excel(wb, data)
        else:
            raise ValueError(f"未対応の取引先: {company_name}")

        # 金額の検証
        validation = validate_totals(wb, data, company_name)

        # 編集済みExcelを保存
        wb.save(output_path)

        # テンプレートからdrawing1.xmlを復元（openpyxlが削除した拡張情報を復元）
        # issue_dateとcompany_nameを渡してTEXT関数のキャッシュ値を設定
        restore_drawing_from_template(template_path, output_path, issue_date, company_name)

        return {
            "success": True,
            "output_path": output_path,
            "validation": validation
        }

    except Exception as e:
        raise RuntimeError(f"Excel編集エラー: {str(e)}") from e


def main():
    """
    メイン関数

    コマンドライン引数からExcel情報を受け取り、編集結果のパスを標準出力に返す。
    """
    if len(sys.argv) != 5:
        print(json.dumps({
            "error": "引数が不足しています",
            "usage": "python3 excel_editor.py <company_name> <template_path> <output_path> <data_json>"
        }, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)

    company_name = sys.argv[1]
    template_path = sys.argv[2]
    output_path = sys.argv[3]
    data_json = sys.argv[4]

    try:
        # JSON文字列をパース
        data = json.loads(data_json)

        # Excel編集
        result = edit_excel(company_name, template_path, output_path, data)

        # 成功時は結果をJSON形式で返す
        print(json.dumps(result, ensure_ascii=False))

    except Exception as e:
        # エラー時はエラー情報をJSON形式で返す
        import traceback
        print(json.dumps({
            "error": str(e),
            "error_type": type(e).__name__,
            "traceback": traceback.format_exc()
        }, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
