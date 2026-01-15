#!/usr/bin/env python3
"""
PDF生成スクリプト（LibreOffice / Excel直接出力 切り替え対応）

Excelファイルの各シートを個別のPDFに変換します。
環境変数 PDF_ENGINE で出力エンジンを切り替えます。

環境変数:
    PDF_ENGINE: 出力エンジンの選択
        - libreoffice（デフォルト）: LibreOfficeのsofficeコマンドを使用
        - excel: Windows側のExcel ExportAsFixedFormatを使用（WSL2経由）

使用法:
    python3 pdf_generator.py <excel_path> <output_dir>

引数:
    excel_path: Excelファイルのパス
    output_dir: 出力ディレクトリのパス

出力:
    生成されたPDFファイルのパス（JSON形式、標準出力）
    {
        "success": true,
        "order_pdf_path": "/tmp/注文書.pdf",
        "inspection_pdf_path": "/tmp/検収書.pdf",
        "engine": "libreoffice" | "excel"
    }

処理ルール（docs/processing_rules.md）:
    - 注文書シートを「注文書_YYMM.pdf」として出力
    - 検収書シートを「検収書_YYMM.pdf」として出力
"""

import sys
import json
import subprocess
import os
import shutil
from pathlib import Path
import openpyxl
from typing import Dict, Any


def check_libreoffice() -> bool:
    """
    LibreOfficeがインストールされているかチェック

    Returns:
        インストール済みならTrue、未インストールならFalse
    """
    try:
        # soffice --version でバージョン情報が取得できるかチェック
        result = subprocess.run(
            ['soffice', '--version'],
            capture_output=True,
            text=True,
            timeout=5
        )
        return result.returncode == 0
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def calculate_formulas_python(wb: openpyxl.Workbook) -> dict:
    """
    Pythonで数式を計算し、計算結果を辞書で返す

    LibreOfficeのheadless変換ではシート間参照の計算に失敗するため、
    Pythonで直接計算を行う。

    Args:
        wb: 編集済みのExcelワークブック

    Returns:
        計算結果の辞書 {シート名: {セル座標: 値}}
    """
    from datetime import datetime
    from calendar import monthrange
    import math

    results = {"注文書": {}, "検収書": {}}

    # 注文書シートから入力値を取得
    ws_order = wb["注文書"]
    ws_inspection = wb["検収書"]

    # AC2: 発行日（入力値）
    issue_date = ws_order['AC2'].value
    if isinstance(issue_date, datetime):
        pass
    elif issue_date is None:
        issue_date = datetime.now().replace(day=1)
    else:
        # 文字列の場合などはdatetimeに変換を試みる
        try:
            issue_date = datetime.strptime(str(issue_date), '%Y-%m-%d')
        except:
            issue_date = datetime.now().replace(day=1)

    # R18, T18: 数量・単価（入力値）
    r18 = ws_order['R18'].value or 0
    t18 = ws_order['T18'].value or 0

    # R20, T20: 検収書の数量・単価（入力値）
    r20 = ws_inspection['R20'].value or 0
    t20 = ws_inspection['T20'].value or 0

    # === 注文書シートの数式計算 ===

    # AC3: =TEXT(AC2,"yyyymmdd")&"-01"
    ac3_value = issue_date.strftime('%Y%m%d') + '-01'
    results["注文書"]["AC3"] = ac3_value

    # C17: =TEXT(注文書!$AC$2,"yyyy年ｍｍ月分作業費")
    c17_value = issue_date.strftime('%Y年%m月分作業費')
    results["注文書"]["C17"] = c17_value

    # AA17: ="見積番号：TRR-"&TEXT(注文書!$AC$2,"yy-")&"0"&TEXT(注文書!$AC$2,"mm")
    yy = issue_date.strftime('%y')
    mm = issue_date.strftime('%m')
    aa17_value = f"見積番号：TRR-{yy}-0{mm}"
    results["注文書"]["AA17"] = aa17_value

    # W17: =T17*R17（この行は空の可能性があるのでスキップ）

    # W18: 明細行の金額（R18 * T18）
    w18_value = r18 * t18
    results["注文書"]["W18"] = w18_value

    # W39: =SUM(W16:Z38) - 簡略化してW18のみを計算（明細1行の場合）
    w39_value = w18_value
    results["注文書"]["W39"] = w39_value

    # W40: =ROUNDDOWN(W39*0.1,0)
    w40_value = math.floor(w39_value * 0.1)
    results["注文書"]["W40"] = w40_value

    # W41: =W39+W40
    w41_value = w39_value + w40_value
    results["注文書"]["W41"] = w41_value

    # G12: =W41
    results["注文書"]["G12"] = w41_value

    # === 検収書シートの数式計算 ===

    # AC4: =注文書!AC3
    results["検収書"]["AC4"] = ac3_value

    # AC5: =EOMONTH(注文書!$AC$2,0) - 月末日
    year = issue_date.year
    month = issue_date.month
    last_day = monthrange(year, month)[1]
    ac5_value = datetime(year, month, last_day)
    results["検収書"]["AC5"] = ac5_value

    # C19: =注文書!C17
    results["検収書"]["C19"] = c17_value

    # AA19: =注文書!AA17
    results["検収書"]["AA19"] = aa17_value

    # W19: =T19*R19（行19は空の可能性があるのでスキップ）

    # W20: 明細行の金額（R20 * T20）
    w20_value = r20 * t20
    results["検収書"]["W20"] = w20_value

    # W41: =SUM(W18:Z40) - 簡略化してW20のみを計算（明細1行の場合）
    w41_inspection = w20_value
    results["検収書"]["W41"] = w41_inspection

    # W42: =ROUNDDOWN(W41*0.1,0)
    w42_value = math.floor(w41_inspection * 0.1)
    results["検収書"]["W42"] = w42_value

    # W43: =W41+W42
    w43_value = w41_inspection + w42_value
    results["検収書"]["W43"] = w43_value

    # G14: =W43
    results["検収書"]["G14"] = w43_value

    return results


def apply_calculated_values(wb: openpyxl.Workbook, calculated_values: dict) -> None:
    """
    計算結果をワークブックに適用する（数式を値に置き換える）

    Args:
        wb: Excelワークブック
        calculated_values: calculate_formulas_pythonの戻り値
    """
    for sheet_name, cells in calculated_values.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for coord, value in cells.items():
                cell = ws[coord]
                # 数式を計算結果で置き換え
                cell.value = value


def resolve_cross_sheet_formulas(wb: openpyxl.Workbook, source_sheet_name: str) -> None:
    """
    シート間参照の数式を、参照元の値で置き換える

    検収書シートは注文書シートを参照する数式を持っているため、
    注文書シートを削除する前に、これらの数式を値に置き換える必要がある。

    Args:
        wb: Excelワークブック
        source_sheet_name: 参照元シート名（削除されるシート）

    Note:
        - 数式内に「シート名!」が含まれるセルを検出
        - 参照元シートの該当セルの値を取得して置き換え
    """
    import re

    source_ws = wb[source_sheet_name] if source_sheet_name in wb.sheetnames else None
    if not source_ws:
        return

    for sheet_name in wb.sheetnames:
        if sheet_name == source_sheet_name:
            continue

        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    formula = str(cell.value)
                    # シート参照を含む数式を検出（例: =注文書!C17, =EOMONTH(注文書!$AC$2,0)）
                    if f'{source_sheet_name}!' in formula:
                        # 参照先の値を取得して置き換え
                        # 単純な参照の場合（例: =注文書!C17）
                        simple_ref_match = re.match(rf'^={source_sheet_name}!\$?([A-Z]+)\$?(\d+)$', formula)
                        if simple_ref_match:
                            col = simple_ref_match.group(1)
                            row_num = int(simple_ref_match.group(2))
                            ref_cell = source_ws[f'{col}{row_num}']
                            cell.value = ref_cell.value
                            cell.data_type = ref_cell.data_type if ref_cell.data_type != 'f' else 'n'


def convert_sheet_to_pdf(excel_path: str, sheet_name: str, output_path: str, calculated_values: dict = None) -> str:
    """
    Excelの特定シートをPDFに変換

    LibreOfficeは特定シートのみのPDF出力をサポートしていないため、
    openpyxlで一時ファイルを作成し、対象シートのみを含むExcelを生成してから変換する。

    Args:
        excel_path: Excelファイルのパス
        sheet_name: 変換対象のシート名
        output_path: 出力PDFファイルのパス
        calculated_values: Pythonで計算した数式の結果（calculate_formulas_pythonの戻り値）

    Returns:
        生成されたPDFファイルのパス

    Raises:
        RuntimeError: LibreOffice未インストールまたは変換失敗時
    """
    # LibreOfficeチェック
    if not check_libreoffice():
        raise RuntimeError(
            "LibreOfficeがインストールされていません。\n"
            "ローカル開発環境ではモック動作となります。\n"
            "本番環境（AWS Lambda Docker Image）ではLibreOfficeを含むイメージを使用してください。"
        )

    # 一時ファイルパスを生成
    output_dir = os.path.dirname(output_path)
    temp_excel_path = os.path.join(output_dir, f"temp_{sheet_name}_{os.getpid()}.xlsx")

    try:
        # 元のExcelを読み込み（数式を保持）
        wb = openpyxl.load_workbook(excel_path)

        # Pythonで計算した値を適用（#NAME?エラー防止）
        # LibreOfficeのheadless変換ではシート間参照の計算に失敗するため、
        # Pythonで計算した値を直接セルに設定する
        if calculated_values:
            apply_calculated_values(wb, calculated_values)

        # 対象シートのみを残す
        sheets_to_remove = [s for s in wb.sheetnames if s != sheet_name]
        for s in sheets_to_remove:
            del wb[s]

        # 印刷スケールを100%に設定
        # LibreOfficeはExcelのスケール設定を正しく解釈できないため、
        # 100%に設定してCubePDFと同じサイズで出力する
        ws = wb[sheet_name]
        ws.page_setup.scale = 100

        # 一時Excelとして保存
        wb.save(temp_excel_path)
        wb.close()

        # LibreOfficeでPDF変換
        result = subprocess.run(
            [
                'soffice',
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', output_dir,
                temp_excel_path
            ],
            capture_output=True,
            text=True,
            timeout=60  # 60秒タイムアウト
        )

        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice変換エラー: {result.stderr}")

        # 生成されたPDFファイルのパスを推測（LibreOfficeは元ファイル名.pdfで出力）
        temp_pdf_path = os.path.join(output_dir, f"temp_{sheet_name}_{os.getpid()}.pdf")

        if not os.path.exists(temp_pdf_path):
            raise RuntimeError(f"PDFファイルが生成されませんでした: {temp_pdf_path}")

        # 出力パスにリネーム
        shutil.move(temp_pdf_path, output_path)

        return output_path

    except subprocess.TimeoutExpired:
        raise RuntimeError("PDF変換がタイムアウトしました（60秒以内に完了しませんでした）")
    except Exception as e:
        raise RuntimeError(f"PDF変換エラー: {str(e)}") from e
    finally:
        # 一時ファイル削除
        if os.path.exists(temp_excel_path):
            os.remove(temp_excel_path)


def get_pdf_engine() -> str:
    """
    環境変数からPDF出力エンジンを取得

    Returns:
        'libreoffice' または 'excel'（デフォルト: 'libreoffice'）
    """
    engine = os.getenv('PDF_ENGINE', 'libreoffice').lower()
    if engine not in ('libreoffice', 'excel'):
        print(f"警告: 不明なPDF_ENGINE値 '{engine}'。デフォルトの'libreoffice'を使用します。", file=sys.stderr)
        return 'libreoffice'
    return engine


def convert_wsl_to_windows_path(wsl_path: str) -> str:
    """
    WSL2パスをWindowsパスに変換

    Args:
        wsl_path: WSL2内のファイルパス

    Returns:
        Windowsパス（例: C:\\Users\\... または \\\\wsl$\\Ubuntu\\...）

    Examples:
        /mnt/c/Users/test.xlsx → C:\\Users\\test.xlsx
        /home/user/file.xlsx → \\\\wsl$\\Ubuntu\\home\\user\\file.xlsx
        /tmp/output.pdf → \\\\wsl$\\Ubuntu\\tmp\\output.pdf
    """
    path = Path(wsl_path).resolve()
    path_str = str(path)

    # /mnt/c/... 形式の場合
    if path_str.startswith('/mnt/'):
        drive = path_str[5].upper()
        rest = path_str[6:].replace('/', '\\')
        return f"{drive}:{rest}"

    # WSL内部パスの場合
    distro = get_wsl_distro_name()
    return f"\\\\wsl$\\{distro}{path_str.replace('/', '\\')}"


def get_wsl_distro_name() -> str:
    """
    WSLディストリビューション名を取得

    Returns:
        ディストリビューション名（取得失敗時は 'Ubuntu'）
    """
    try:
        result = subprocess.run(
            ['wslpath', '-w', '/'],
            capture_output=True,
            text=True,
            timeout=5
        )
        if result.returncode == 0:
            # \\wsl$\Ubuntu\ のような形式から抽出
            win_path = result.stdout.strip()
            parts = win_path.split('\\')
            for i, part in enumerate(parts):
                if part == 'wsl$' and i + 1 < len(parts):
                    return parts[i + 1]
    except Exception:
        pass
    return 'Ubuntu'  # デフォルト


def generate_excel_export_script(excel_path: str, order_pdf_path: str, inspection_pdf_path: str) -> str:
    """
    ExcelのExportAsFixedFormatでPDF出力するPowerShellスクリプトを生成

    Args:
        excel_path: Excelファイルのパス（Windowsパス形式）
        order_pdf_path: 注文書PDFの出力パス（Windowsパス形式）
        inspection_pdf_path: 検収書PDFの出力パス（Windowsパス形式）

    Returns:
        PowerShellスクリプト文字列

    Note:
        - CubePDFはGUI非表示での実行が困難なため、Excel標準機能を使用
        - xlTypePDF = 0 でPDF形式を指定
        - シート別にExportAsFixedFormatを呼び出し
    """
    # パス内のバックスラッシュをエスケープ
    excel_path_escaped = excel_path.replace('\\', '\\\\')
    order_pdf_escaped = order_pdf_path.replace('\\', '\\\\')
    inspection_pdf_escaped = inspection_pdf_path.replace('\\', '\\\\')

    script = f'''
$ErrorActionPreference = "Stop"

$excel = $null
$workbook = $null

try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false

    $workbook = $excel.Workbooks.Open("{excel_path_escaped}")

    # 注文書シートをPDF出力
    $orderSheet = $workbook.Worksheets.Item(1)
    $orderSheet.ExportAsFixedFormat(0, "{order_pdf_escaped}")

    # 検収書シートをPDF出力
    $inspectionSheet = $workbook.Worksheets.Item(2)
    $inspectionSheet.ExportAsFixedFormat(0, "{inspection_pdf_escaped}")

    $workbook.Close($false)
    $excel.Quit()

    Write-Output "SUCCESS"
}} catch {{
    Write-Error $_.Exception.Message
    exit 1
}} finally {{
    if ($workbook -ne $null) {{
        [System.Runtime.Interopservices.Marshal]::ReleaseComObject($workbook) | Out-Null
    }}
    if ($excel -ne $null) {{
        [System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null
    }}
    [System.GC]::Collect()
    [System.GC]::WaitForPendingFinalizers()
}}
'''
    return script


def check_excel_available() -> bool:
    """
    Windows側でExcelが利用可能かチェック（WSL2環境用）

    Returns:
        利用可能ならTrue、不可ならFalse
    """
    try:
        # PowerShellでExcel COMオブジェクトが作成できるかチェック
        result = subprocess.run(
            ['powershell.exe', '-ExecutionPolicy', 'Bypass', '-Command',
             '$excel = New-Object -ComObject Excel.Application; $excel.Quit(); Write-Output "OK"'],
            capture_output=True,
            text=True,
            timeout=30
        )
        return result.returncode == 0 and 'OK' in result.stdout
    except Exception:
        return False


def convert_excel_sheets_to_pdf_excel(excel_path: str, output_dir: str) -> Dict[str, str]:
    """
    Excel ExportAsFixedFormatでPDF変換（WSL2からWindows呼び出し）

    Args:
        excel_path: Excelファイルのパス
        output_dir: 出力ディレクトリのパス

    Returns:
        生成されたPDFファイルのパス辞書

    Raises:
        RuntimeError: 変換失敗時
    """
    # 出力ファイルパス
    order_pdf_path = os.path.join(output_dir, f"order_{os.getpid()}.pdf")
    inspection_pdf_path = os.path.join(output_dir, f"inspection_{os.getpid()}.pdf")

    try:
        # WSLパス → Windowsパス変換
        win_excel_path = convert_wsl_to_windows_path(excel_path)
        win_order_pdf_path = convert_wsl_to_windows_path(order_pdf_path)
        win_inspection_pdf_path = convert_wsl_to_windows_path(inspection_pdf_path)

        # PowerShellスクリプト生成
        ps_script = generate_excel_export_script(
            win_excel_path,
            win_order_pdf_path,
            win_inspection_pdf_path
        )

        # PowerShell実行
        # 日本語Windowsの場合、出力はcp932でエンコードされるため、
        # text=Falseでバイト列として受け取り、手動でデコードする
        result = subprocess.run(
            ['powershell.exe', '-ExecutionPolicy', 'Bypass', '-Command', ps_script],
            capture_output=True,
            text=False,
            timeout=120  # 2分タイムアウト
        )

        if result.returncode != 0:
            # エラーメッセージをデコード（cp932 → UTF-8、失敗時は置換）
            try:
                error_msg = result.stderr.decode('cp932', errors='replace').strip() if result.stderr else "不明なエラー"
            except Exception:
                error_msg = result.stderr.decode('utf-8', errors='replace').strip() if result.stderr else "不明なエラー"
            raise RuntimeError(f"Excel PDF出力エラー: {error_msg}")

        # 出力ファイルの存在確認
        if not os.path.exists(order_pdf_path):
            raise RuntimeError(f"注文書PDFが生成されませんでした: {order_pdf_path}")
        if not os.path.exists(inspection_pdf_path):
            raise RuntimeError(f"検収書PDFが生成されませんでした: {inspection_pdf_path}")

        return {
            "order_pdf_path": order_pdf_path,
            "inspection_pdf_path": inspection_pdf_path
        }

    except subprocess.TimeoutExpired:
        raise RuntimeError("PDF変換がタイムアウトしました（120秒以内に完了しませんでした）")
    except Exception as e:
        raise RuntimeError(f"Excel PDF出力エラー: {str(e)}") from e


def convert_excel_sheets_to_pdf_libreoffice(excel_path: str, output_dir: str) -> Dict[str, str]:
    """
    LibreOfficeでExcel→PDF変換（既存処理）

    Args:
        excel_path: Excelファイルのパス
        output_dir: 出力ディレクトリのパス

    Returns:
        生成されたPDFファイルのパス辞書

    Raises:
        RuntimeError: 変換失敗時

    Note:
        openpyxlを経由するとExcel XMLのキャッシュ値が失われ、LibreOfficeがTEXT関数を
        正しく計算できない問題があった。そのため、元のExcelを直接LibreOfficeでPDF変換し、
        pypdfでページを分割する方式に変更した。
    """
    from pypdf import PdfReader, PdfWriter

    # LibreOfficeチェック
    if not check_libreoffice():
        raise RuntimeError(
            "LibreOfficeがインストールされていません。\n"
            "本番環境（AWS Lambda Docker Image）ではLibreOfficeを含むイメージを使用してください。"
        )

    # 一時PDFパス
    temp_full_pdf_path = os.path.join(output_dir, f"temp_full_{os.getpid()}.pdf")

    try:
        # LibreOfficeで全シートをPDFに変換（openpyxlを経由しない）
        # これによりexcel_editor.pyで設定したキャッシュ値が保持される
        result = subprocess.run(
            [
                'soffice',
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', output_dir,
                excel_path
            ],
            capture_output=True,
            text=True,
            timeout=60
        )

        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice変換エラー: {result.stderr}")

        # LibreOfficeが生成するPDFのファイル名を推測
        excel_filename = os.path.basename(excel_path)
        generated_pdf_name = os.path.splitext(excel_filename)[0] + '.pdf'
        generated_pdf_path = os.path.join(output_dir, generated_pdf_name)

        if not os.path.exists(generated_pdf_path):
            raise RuntimeError(f"PDFファイルが生成されませんでした: {generated_pdf_path}")

        # 一時ファイル名にリネーム
        shutil.move(generated_pdf_path, temp_full_pdf_path)

        # PDFをページごとに分割（注文書=1ページ目、検収書=2ページ目）
        reader = PdfReader(temp_full_pdf_path)

        if len(reader.pages) < 2:
            raise RuntimeError(f"PDFのページ数が不足しています: {len(reader.pages)}ページ")

        # 注文書（1ページ目）
        order_pdf_path = os.path.join(output_dir, f"order_{os.getpid()}.pdf")
        order_writer = PdfWriter()
        order_writer.add_page(reader.pages[0])
        with open(order_pdf_path, 'wb') as f:
            order_writer.write(f)

        # 検収書（2ページ目）
        inspection_pdf_path = os.path.join(output_dir, f"inspection_{os.getpid()}.pdf")
        inspection_writer = PdfWriter()
        inspection_writer.add_page(reader.pages[1])
        with open(inspection_pdf_path, 'wb') as f:
            inspection_writer.write(f)

        return {
            "order_pdf_path": order_pdf_path,
            "inspection_pdf_path": inspection_pdf_path
        }

    except subprocess.TimeoutExpired:
        raise RuntimeError("PDF変換がタイムアウトしました（60秒以内に完了しませんでした）")
    except Exception as e:
        raise RuntimeError(f"PDF変換エラー: {str(e)}") from e
    finally:
        # 一時ファイル削除
        if os.path.exists(temp_full_pdf_path):
            os.remove(temp_full_pdf_path)


def convert_excel_sheets_to_pdf(excel_path: str, output_dir: str) -> Dict[str, str]:
    """
    Excelファイルの注文書シートと検収書シートをそれぞれPDFに変換

    環境変数 PDF_ENGINE に応じてエンジンを切り替え:
        - libreoffice（デフォルト）: LibreOfficeを使用
        - excel: Windows側のExcel ExportAsFixedFormatを使用

    Args:
        excel_path: Excelファイルのパス
        output_dir: 出力ディレクトリのパス

    Returns:
        生成されたPDFファイルのパス辞書
        {
            "order_pdf_path": "/tmp/order.pdf",
            "inspection_pdf_path": "/tmp/inspection.pdf"
        }

    Raises:
        RuntimeError: 変換失敗時
    """
    engine = get_pdf_engine()

    if engine == 'excel':
        return convert_excel_sheets_to_pdf_excel(excel_path, output_dir)
    else:
        return convert_excel_sheets_to_pdf_libreoffice(excel_path, output_dir)


def main():
    """
    メイン関数

    コマンドライン引数からExcelパスを受け取り、PDF変換結果のパスを標準出力に返す。
    """
    if len(sys.argv) != 3:
        print(json.dumps({
            "error": "引数が不足しています",
            "usage": "python3 pdf_generator.py <excel_path> <output_dir>"
        }, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)

    excel_path = sys.argv[1]
    output_dir = sys.argv[2]

    try:
        # 使用エンジンを取得
        engine = get_pdf_engine()

        # PDF生成（注文書・検収書シートをそれぞれPDFに変換）
        result = convert_excel_sheets_to_pdf(excel_path, output_dir)

        # 成功時は出力パスをJSON形式で返す
        print(json.dumps({
            "success": True,
            "order_pdf_path": result["order_pdf_path"],
            "inspection_pdf_path": result["inspection_pdf_path"],
            "engine": engine
        }, ensure_ascii=False))

    except Exception as e:
        # エラー時はエラー情報をJSON形式で返す
        import traceback
        print(json.dumps({
            "error": str(e),
            "error_type": type(e).__name__,
            "traceback": traceback.format_exc(),
            "engine": get_pdf_engine()
        }, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
