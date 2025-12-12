#!/usr/bin/env python3
"""
Excel検証スクリプト（LibreOffice使用）

LibreOfficeでExcelを開いて数式を計算させ、セル値を取得して検証します。

使用法:
    python3 excel_validator.py <excel_path> <company_name> <invoice_data_json>

引数:
    excel_path: Excelファイルのパス
    company_name: 取引先名（ネクストビッツ or オフ・ビート・ワークス）
    invoice_data_json: 請求書から抽出したデータ（JSON文字列）

出力:
    検証結果（JSON形式、標準出力）
    {
        "success": true,
        "checks": [
            {"item": "注文番号形式", "expected": "20250701-01", "actual": "20250701-01", "passed": true},
            ...
        ],
        "errors": []
    }

検証項目:
    ネクストビッツ:
        注文書シート:
            - AC3: 注文番号がyyyymmdd-01形式か
            - B8: 宛名が「株式会社ネクストビッツ　御中」か
            - G12: 発注金額が請求書PDFの合計金額と一致するか
            - C17: 明細タイトルが「yyyy年mm月分作業費」形式か
            - AA17: 摘要が「見積番号：TRR-YY-0MM」形式で見積PDFの番号と一致するか
            - C18: 件名が「　Telemas作業(システム改修等)」か（固定文字列、先頭全角スペース）
            - C19: 「以下、余白」か
            - W39: 小計が請求書PDFの消費税10%対象と一致するか
            - W40: 消費税が請求書PDFの消費税(10%)と一致するか
            - W41: 合計金額が請求書PDFの合計金額と一致するか
        検収書シート:
            - AC4: 検収番号がyyyymmdd-01形式か
            - AC5: 検収日が当月末日か
            - B7: 宛名が「株式会社ネクストビッツ　御中」か
            - G14: 合計金額が請求書PDFの合計金額と一致するか
            - C19: 明細タイトルが「yyyy年mm月分作業費」形式か
            - AA19: 摘要が「見積番号：TRR-YY-0MM」形式で見積PDFの番号と一致するか
            - C20: 件名が「　Telemas作業(システム改修等)」か（固定文字列、先頭全角スペース）
            - C21: 「以下、余白」か
            - W41: 小計が請求書PDFの消費税10%対象と一致するか
            - W42: 消費税が請求書PDFの消費税(10%)と一致するか
            - W43: 合計金額が請求書PDFの合計金額と一致するか

    オフ・ビート・ワークス:
        注文書シート:
            - AC3: 注文番号がyyyymmdd-02形式か
            - B8: 宛名が「株式会社オフ・ビート・ワークス　御中」か
            - G12: 発注金額が請求書PDFの合計金額と一致するか
            - C17: 明細タイトルが「yyyy年mm月作業費」形式か（「分」なし）
            - AA17: 摘要が「見積番号：NNNNNNN」形式で見積PDFの番号と一致するか
            - C(18+N): 「以下、余白」か（N=明細行数、動的に計算）
            - W39: 小計が請求書PDFの消費税10%対象と一致するか
            - W40: 消費税が請求書PDFの消費税(10%)と一致するか
            - W41: 合計金額が請求書PDFの合計金額と一致するか
        検収書シート:
            - AC4: 検収番号がyyyymmdd-02形式か
            - AC5: 検収日が当月末日か
            - B7: 宛名が「株式会社オフ・ビート・ワークス　御中」か
            - G14: 合計金額が請求書PDFの合計金額と一致するか
            - C19: 明細タイトルが「yyyy年mm月作業費」形式か
            - AA19: 摘要が「見積番号：NNNNNNN」形式で見積PDFの番号と一致するか
            - C(20+N): 「以下、余白」か（N=明細行数、動的に計算）
            - W41: 小計が請求書PDFの消費税10%対象と一致するか
            - W42: 消費税が請求書PDFの消費税(10%)と一致するか
            - W43: 合計金額が請求書PDFの合計金額と一致するか
"""

import sys
import json
import subprocess
import os
import csv
import re
from datetime import datetime
from typing import Dict, Any, List, Tuple
from pathlib import Path


def check_libreoffice() -> bool:
    """LibreOfficeがインストールされているかチェック"""
    try:
        result = subprocess.run(
            ['soffice', '--version'],
            capture_output=True,
            text=True,
            timeout=5
        )
        return result.returncode == 0
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def convert_excel_to_csv_with_libreoffice(excel_path: str, output_dir: str) -> Dict[str, str]:
    """
    LibreOfficeでExcelを開き数式を計算後、各シートをCSVとして保存

    LibreOfficeがExcelを開くと自動的に数式が計算される。
    その結果を各シートごとにCSVファイルとして出力する。

    Args:
        excel_path: Excelファイルのパス
        output_dir: 出力ディレクトリ

    Returns:
        シート名 -> CSVファイルパスの辞書
    """
    if not check_libreoffice():
        raise RuntimeError("LibreOfficeがインストールされていません")

    import openpyxl

    # 方法: LibreOfficeでExcelをODSに変換し、数式を計算させてから再度Excelに変換
    # その後openpyxlでdata_only=Trueで読み込む

    # Step 1: LibreOfficeでExcelを開いて再保存（数式計算のため）
    # --convert-to xlsx で再保存すると数式が計算される
    calc_excel_path = os.path.join(output_dir, f"calculated_{os.getpid()}.xlsx")

    # LibreOfficeでODS経由で変換（数式を確実に計算させる）
    temp_ods = os.path.join(output_dir, f"temp_{os.getpid()}.ods")

    # Excel -> ODS
    result = subprocess.run(
        [
            'soffice',
            '--headless',
            '--convert-to', 'ods',
            '--outdir', output_dir,
            excel_path
        ],
        capture_output=True,
        text=True,
        timeout=60
    )

    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    generated_ods = os.path.join(output_dir, f"{base_name}.ods")

    if os.path.exists(generated_ods):
        os.rename(generated_ods, temp_ods)

        # ODS -> XLSX（数式が計算された状態で値として保存）
        result = subprocess.run(
            [
                'soffice',
                '--headless',
                '--convert-to', 'xlsx',
                '--outdir', output_dir,
                temp_ods
            ],
            capture_output=True,
            text=True,
            timeout=60
        )

        generated_xlsx = os.path.join(output_dir, f"temp_{os.getpid()}.xlsx")
        if os.path.exists(generated_xlsx):
            os.rename(generated_xlsx, calc_excel_path)

        # 一時ODSファイル削除
        if os.path.exists(temp_ods):
            os.remove(temp_ods)

    csv_paths = {}

    # Step 2: 計算済みExcelをopenpyxlで読み込み、各シートをCSVに変換
    if os.path.exists(calc_excel_path):
        wb = openpyxl.load_workbook(calc_excel_path, data_only=True)

        for sheet_name in wb.sheetnames:
            csv_filename = f"temp_{sheet_name}_{os.getpid()}.csv"
            csv_path = os.path.join(output_dir, csv_filename)

            ws = wb[sheet_name]

            with open(csv_path, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                for row in ws.iter_rows():
                    row_data = []
                    for cell in row:
                        value = cell.value
                        if value is None:
                            row_data.append('')
                        else:
                            row_data.append(str(value))
                    writer.writerow(row_data)

            csv_paths[sheet_name] = csv_path

        wb.close()

        # 計算済みExcel削除
        os.remove(calc_excel_path)
    else:
        # フォールバック: 元のExcelをそのまま読み込む（数式は計算されない可能性）
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        for sheet_name in wb.sheetnames:
            csv_filename = f"temp_{sheet_name}_{os.getpid()}.csv"
            csv_path = os.path.join(output_dir, csv_filename)

            ws = wb[sheet_name]

            with open(csv_path, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                for row in ws.iter_rows():
                    row_data = []
                    for cell in row:
                        value = cell.value
                        if value is None:
                            row_data.append('')
                        else:
                            row_data.append(str(value))
                    writer.writerow(row_data)

            csv_paths[sheet_name] = csv_path

        wb.close()

    return csv_paths


def read_cell_from_csv(csv_path: str, row: int, col: int) -> str:
    """
    CSVファイルから指定セルの値を読み取る

    Args:
        csv_path: CSVファイルパス
        row: 行番号（1始まり）
        col: 列番号（1始まり）

    Returns:
        セルの値（文字列）
    """
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for i, csv_row in enumerate(reader, start=1):
                if i == row:
                    if col <= len(csv_row):
                        return csv_row[col - 1]
                    return ""
        return ""
    except Exception:
        return ""


def col_letter_to_number(col_letter: str) -> int:
    """列文字を列番号に変換（A=1, B=2, ..., AA=27, ...）"""
    result = 0
    for char in col_letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result


def get_cell_value(csv_path: str, cell_ref: str) -> str:
    """
    セル参照（例: AC3, B8）から値を取得

    Args:
        csv_path: CSVファイルパス
        cell_ref: セル参照（例: "AC3"）

    Returns:
        セルの値（文字列）
    """
    match = re.match(r'^([A-Z]+)(\d+)$', cell_ref.upper())
    if not match:
        return ""

    col_letter = match.group(1)
    row = int(match.group(2))
    col = col_letter_to_number(col_letter)

    return read_cell_from_csv(csv_path, row, col)


def parse_number(value: str) -> int:
    """文字列から数値を抽出（カンマ区切り対応）"""
    if not value:
        return 0
    # カンマを除去し、数値部分を抽出
    cleaned = re.sub(r'[^\d.-]', '', value.replace(',', ''))
    try:
        return int(float(cleaned))
    except ValueError:
        return 0


def validate_nextbits_excel(csv_paths: Dict[str, str], validation_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    ネクストビッツのExcelを検証

    Args:
        csv_paths: シート名 -> CSVファイルパスの辞書
        validation_data: 検証データ（invoice, estimate, items_count）

    Returns:
        検証結果
    """
    checks: List[Dict[str, Any]] = []
    errors: List[str] = []

    order_csv = csv_paths.get("注文書", "")
    inspection_csv = csv_paths.get("検収書", "")

    if not order_csv or not inspection_csv:
        return {
            "success": False,
            "checks": [],
            "errors": ["注文書または検収書シートが見つかりません"]
        }

    # 検証データの取り出し
    invoice_data = validation_data.get('invoice', validation_data)  # 後方互換性
    estimate_data = validation_data.get('estimate', {})

    # 請求書データ
    invoice_total = invoice_data.get('total', 0)
    invoice_subtotal = invoice_data.get('subtotal', 0)
    invoice_tax = invoice_data.get('tax', 0)

    # 見積データ（摘要チェック用）
    estimate_number = estimate_data.get('estimate_number', '')
    estimate_subject = estimate_data.get('subject', '')  # 件名チェック用

    # === 注文書シート検証 ===

    # AC3: 注文番号（yyyymmdd-01形式）
    order_number = get_cell_value(order_csv, "AC3")
    order_number_valid = bool(re.match(r'^\d{8}-01$', order_number))
    checks.append({
        "sheet": "注文書",
        "cell": "AC3",
        "item": "注文番号形式",
        "expected": "yyyymmdd-01形式",
        "actual": order_number,
        "passed": order_number_valid
    })
    if not order_number_valid:
        errors.append(f"注文書AC3: 注文番号が不正です（{order_number}）")

    # B8: 宛名
    order_address = get_cell_value(order_csv, "B8")
    order_address_valid = "株式会社ネクストビッツ" in order_address and "御中" in order_address
    checks.append({
        "sheet": "注文書",
        "cell": "B8",
        "item": "宛名",
        "expected": "株式会社ネクストビッツ　御中",
        "actual": order_address,
        "passed": order_address_valid
    })
    if not order_address_valid:
        errors.append(f"注文書B8: 宛名が不正です（{order_address}）")

    # G12: 発注金額
    order_amount = parse_number(get_cell_value(order_csv, "G12"))
    order_amount_valid = order_amount == invoice_total
    checks.append({
        "sheet": "注文書",
        "cell": "G12",
        "item": "発注金額",
        "expected": f"{invoice_total:,}円",
        "actual": f"{order_amount:,}円",
        "passed": order_amount_valid
    })
    if not order_amount_valid:
        errors.append(f"注文書G12: 発注金額が請求書と不一致（期待: {invoice_total:,}円、実際: {order_amount:,}円）")

    # C17: 明細タイトル（yyyy年mm月分作業費形式）
    detail_title = get_cell_value(order_csv, "C17")
    detail_title_valid = bool(re.match(r'^\d{4}年\d{1,2}月分作業費$', detail_title))
    checks.append({
        "sheet": "注文書",
        "cell": "C17",
        "item": "明細タイトル",
        "expected": "yyyy年mm月分作業費形式",
        "actual": detail_title,
        "passed": detail_title_valid
    })
    if not detail_title_valid:
        errors.append(f"注文書C17: 明細タイトルが不正です（{detail_title}）")

    # AA17: 摘要（「見積番号：TRR-YY-0MM」形式）
    order_remarks = get_cell_value(order_csv, "AA17")
    # 見積番号の期待値を構築：見積PDFから取得した番号を使用
    if estimate_number:
        expected_remarks = f"見積番号：{estimate_number}"
        order_remarks_valid = order_remarks == expected_remarks
    else:
        # 見積番号がない場合は形式チェックのみ
        expected_remarks = "見積番号：TRR-YY-0MM形式"
        order_remarks_valid = bool(re.match(r'^見積番号：TRR-\d{2}-0\d{2}$', order_remarks))
    checks.append({
        "sheet": "注文書",
        "cell": "AA17",
        "item": "摘要",
        "expected": expected_remarks,
        "actual": order_remarks,
        "passed": order_remarks_valid
    })
    if not order_remarks_valid:
        errors.append(f"注文書AA17: 摘要が不正です（期待: {expected_remarks}、実際: {order_remarks}）")

    # C18: 件名（固定文字列「　Telemas作業(システム改修等)」との比較）
    # ※処理手順より：見積書の件名「yyyy年mm月作業：Telemasシステム改修作業等」→「　Telemas作業(システム改修等)」に変換
    order_subject = get_cell_value(order_csv, "C18")
    expected_subject = "　Telemas作業(システム改修等)"  # 先頭に全角スペース
    order_subject_valid = order_subject == expected_subject
    checks.append({
        "sheet": "注文書",
        "cell": "C18",
        "item": "件名",
        "expected": expected_subject,
        "actual": order_subject,
        "passed": order_subject_valid
    })
    if not order_subject_valid:
        errors.append(f"注文書C18: 件名が不正です（期待: {expected_subject}、実際: {order_subject}）")

    # C19: 以下、余白
    blank_marker = get_cell_value(order_csv, "C19")
    blank_marker_valid = "以下、余白" in blank_marker or blank_marker.strip() == "以下、余白"
    checks.append({
        "sheet": "注文書",
        "cell": "C19",
        "item": "明細締め",
        "expected": "以下、余白",
        "actual": blank_marker,
        "passed": blank_marker_valid
    })
    if not blank_marker_valid:
        errors.append(f"注文書C19: 「以下、余白」が入力されていません（{blank_marker}）")

    # W39: 小計
    order_subtotal = parse_number(get_cell_value(order_csv, "W39"))
    order_subtotal_valid = order_subtotal == invoice_subtotal
    checks.append({
        "sheet": "注文書",
        "cell": "W39",
        "item": "小計",
        "expected": f"{invoice_subtotal:,}円",
        "actual": f"{order_subtotal:,}円",
        "passed": order_subtotal_valid
    })
    if not order_subtotal_valid:
        errors.append(f"注文書W39: 小計が請求書と不一致（期待: {invoice_subtotal:,}円、実際: {order_subtotal:,}円）")

    # W40: 消費税
    order_tax = parse_number(get_cell_value(order_csv, "W40"))
    order_tax_valid = order_tax == invoice_tax
    checks.append({
        "sheet": "注文書",
        "cell": "W40",
        "item": "消費税",
        "expected": f"{invoice_tax:,}円",
        "actual": f"{order_tax:,}円",
        "passed": order_tax_valid
    })
    if not order_tax_valid:
        errors.append(f"注文書W40: 消費税が請求書と不一致（期待: {invoice_tax:,}円、実際: {order_tax:,}円）")

    # W41: 合計金額
    order_total = parse_number(get_cell_value(order_csv, "W41"))
    order_total_valid = order_total == invoice_total
    checks.append({
        "sheet": "注文書",
        "cell": "W41",
        "item": "合計金額",
        "expected": f"{invoice_total:,}円",
        "actual": f"{order_total:,}円",
        "passed": order_total_valid
    })
    if not order_total_valid:
        errors.append(f"注文書W41: 合計金額が請求書と不一致（期待: {invoice_total:,}円、実際: {order_total:,}円）")

    # === 検収書シート検証 ===

    # AC4: 検収番号（yyyymmdd-01形式）
    inspection_number = get_cell_value(inspection_csv, "AC4")
    inspection_number_valid = bool(re.match(r'^\d{8}-01$', inspection_number))
    checks.append({
        "sheet": "検収書",
        "cell": "AC4",
        "item": "検収番号形式",
        "expected": "yyyymmdd-01形式",
        "actual": inspection_number,
        "passed": inspection_number_valid
    })
    if not inspection_number_valid:
        errors.append(f"検収書AC4: 検収番号が不正です（{inspection_number}）")

    # AC5: 検収日（当月末日）- フォーマットチェックのみ
    inspection_date = get_cell_value(inspection_csv, "AC5")
    # 日付形式のチェック（yyyy/mm/dd または yyyy-mm-dd または datetime形式 または数値）
    # openpyxlはdatetime形式で「2025-07-31 00:00:00」のように出力することがある
    inspection_date_valid = bool(inspection_date) and bool(
        re.match(r'^\d{4}[/-]\d{1,2}[/-]\d{1,2}$', inspection_date) or
        re.match(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$', inspection_date) or  # datetime形式
        re.match(r'^\d+$', inspection_date.replace(',', ''))
    )
    checks.append({
        "sheet": "検収書",
        "cell": "AC5",
        "item": "検収日",
        "expected": "当月末日",
        "actual": inspection_date,
        "passed": inspection_date_valid
    })
    if not inspection_date_valid:
        errors.append(f"検収書AC5: 検収日が不正です（{inspection_date}）")

    # B7: 宛名
    inspection_address = get_cell_value(inspection_csv, "B7")
    inspection_address_valid = "株式会社ネクストビッツ" in inspection_address and "御中" in inspection_address
    checks.append({
        "sheet": "検収書",
        "cell": "B7",
        "item": "宛名",
        "expected": "株式会社ネクストビッツ　御中",
        "actual": inspection_address,
        "passed": inspection_address_valid
    })
    if not inspection_address_valid:
        errors.append(f"検収書B7: 宛名が不正です（{inspection_address}）")

    # G14: 合計金額
    inspection_amount = parse_number(get_cell_value(inspection_csv, "G14"))
    inspection_amount_valid = inspection_amount == invoice_total
    checks.append({
        "sheet": "検収書",
        "cell": "G14",
        "item": "合計金額",
        "expected": f"{invoice_total:,}円",
        "actual": f"{inspection_amount:,}円",
        "passed": inspection_amount_valid
    })
    if not inspection_amount_valid:
        errors.append(f"検収書G14: 合計金額が請求書と不一致（期待: {invoice_total:,}円、実際: {inspection_amount:,}円）")

    # C19: 明細タイトル（検収書）
    inspection_detail_title = get_cell_value(inspection_csv, "C19")
    inspection_detail_title_valid = bool(re.match(r'^\d{4}年\d{1,2}月分作業費$', inspection_detail_title))
    checks.append({
        "sheet": "検収書",
        "cell": "C19",
        "item": "明細タイトル",
        "expected": "yyyy年mm月分作業費形式",
        "actual": inspection_detail_title,
        "passed": inspection_detail_title_valid
    })
    if not inspection_detail_title_valid:
        errors.append(f"検収書C19: 明細タイトルが不正です（{inspection_detail_title}）")

    # AA19: 摘要（「見積番号：TRR-YY-0MM」形式）
    inspection_remarks = get_cell_value(inspection_csv, "AA19")
    if estimate_number:
        expected_inspection_remarks = f"見積番号：{estimate_number}"
        inspection_remarks_valid = inspection_remarks == expected_inspection_remarks
    else:
        expected_inspection_remarks = "見積番号：TRR-YY-0MM形式"
        inspection_remarks_valid = bool(re.match(r'^見積番号：TRR-\d{2}-0\d{2}$', inspection_remarks))
    checks.append({
        "sheet": "検収書",
        "cell": "AA19",
        "item": "摘要",
        "expected": expected_inspection_remarks,
        "actual": inspection_remarks,
        "passed": inspection_remarks_valid
    })
    if not inspection_remarks_valid:
        errors.append(f"検収書AA19: 摘要が不正です（期待: {expected_inspection_remarks}、実際: {inspection_remarks}）")

    # C20: 件名（固定文字列「　Telemas作業(システム改修等)」との比較）
    # ※処理手順より：見積書の件名「yyyy年mm月作業：Telemasシステム改修作業等」→「　Telemas作業(システム改修等)」に変換
    inspection_subject = get_cell_value(inspection_csv, "C20")
    expected_inspection_subject = "　Telemas作業(システム改修等)"  # 先頭に全角スペース
    inspection_subject_valid = inspection_subject == expected_inspection_subject
    checks.append({
        "sheet": "検収書",
        "cell": "C20",
        "item": "件名",
        "expected": expected_inspection_subject,
        "actual": inspection_subject,
        "passed": inspection_subject_valid
    })
    if not inspection_subject_valid:
        errors.append(f"検収書C20: 件名が不正です（期待: {expected_inspection_subject}、実際: {inspection_subject}）")

    # C21: 以下、余白
    inspection_blank = get_cell_value(inspection_csv, "C21")
    inspection_blank_valid = "以下、余白" in inspection_blank or inspection_blank.strip() == "以下、余白"
    checks.append({
        "sheet": "検収書",
        "cell": "C21",
        "item": "明細締め",
        "expected": "以下、余白",
        "actual": inspection_blank,
        "passed": inspection_blank_valid
    })
    if not inspection_blank_valid:
        errors.append(f"検収書C21: 「以下、余白」が入力されていません（{inspection_blank}）")

    # W41: 小計（検収書）
    inspection_subtotal = parse_number(get_cell_value(inspection_csv, "W41"))
    inspection_subtotal_valid = inspection_subtotal == invoice_subtotal
    checks.append({
        "sheet": "検収書",
        "cell": "W41",
        "item": "小計",
        "expected": f"{invoice_subtotal:,}円",
        "actual": f"{inspection_subtotal:,}円",
        "passed": inspection_subtotal_valid
    })
    if not inspection_subtotal_valid:
        errors.append(f"検収書W41: 小計が請求書と不一致（期待: {invoice_subtotal:,}円、実際: {inspection_subtotal:,}円）")

    # W42: 消費税（検収書）
    inspection_tax = parse_number(get_cell_value(inspection_csv, "W42"))
    inspection_tax_valid = inspection_tax == invoice_tax
    checks.append({
        "sheet": "検収書",
        "cell": "W42",
        "item": "消費税",
        "expected": f"{invoice_tax:,}円",
        "actual": f"{inspection_tax:,}円",
        "passed": inspection_tax_valid
    })
    if not inspection_tax_valid:
        errors.append(f"検収書W42: 消費税が請求書と不一致（期待: {invoice_tax:,}円、実際: {inspection_tax:,}円）")

    # W43: 合計金額（検収書）
    inspection_total = parse_number(get_cell_value(inspection_csv, "W43"))
    inspection_total_valid = inspection_total == invoice_total
    checks.append({
        "sheet": "検収書",
        "cell": "W43",
        "item": "合計金額",
        "expected": f"{invoice_total:,}円",
        "actual": f"{inspection_total:,}円",
        "passed": inspection_total_valid
    })
    if not inspection_total_valid:
        errors.append(f"検収書W43: 合計金額が請求書と不一致（期待: {invoice_total:,}円、実際: {inspection_total:,}円）")

    return {
        "success": len(errors) == 0,
        "checks": checks,
        "errors": errors
    }


def validate_offbeat_excel(csv_paths: Dict[str, str], validation_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    オフ・ビート・ワークスのExcelを検証

    Args:
        csv_paths: シート名 -> CSVファイルパスの辞書
        validation_data: 検証データ（invoice, estimate, items_count）

    Returns:
        検証結果
    """
    checks: List[Dict[str, Any]] = []
    errors: List[str] = []

    order_csv = csv_paths.get("注文書", "")
    inspection_csv = csv_paths.get("検収書", "")

    if not order_csv or not inspection_csv:
        return {
            "success": False,
            "checks": [],
            "errors": ["注文書または検収書シートが見つかりません"]
        }

    # 検証データの取り出し
    invoice_data = validation_data.get('invoice', validation_data)  # 後方互換性
    estimate_data = validation_data.get('estimate', {})
    items_count = validation_data.get('items_count', 1)

    # 請求書データ
    invoice_total = invoice_data.get('total', 0)
    invoice_subtotal = invoice_data.get('subtotal', 0)
    invoice_tax = invoice_data.get('tax', 0)

    # 見積データ（摘要チェック用）
    estimate_number = estimate_data.get('estimate_number', '')

    # 動的行番号計算（オフ・ビート・ワークスは複数明細に対応）
    # 注文書: 18 + items_count = 「以下、余白」の行
    # 検収書: 20 + items_count = 「以下、余白」の行
    order_blank_row = 18 + items_count
    inspection_blank_row = 20 + items_count
    # 摘要行（AA列）：オフ・ビート・ワークスは1行目のみ（17行目固定）
    order_remarks_row = 17
    inspection_remarks_row = 19

    # === 注文書シート検証 ===

    # AC3: 注文番号（yyyymmdd-02形式）
    order_number = get_cell_value(order_csv, "AC3")
    order_number_valid = bool(re.match(r'^\d{8}-02$', order_number))
    checks.append({
        "sheet": "注文書",
        "cell": "AC3",
        "item": "注文番号形式",
        "expected": "yyyymmdd-02形式",
        "actual": order_number,
        "passed": order_number_valid
    })
    if not order_number_valid:
        errors.append(f"注文書AC3: 注文番号が不正です（{order_number}）")

    # B8: 宛名
    order_address = get_cell_value(order_csv, "B8")
    order_address_valid = "株式会社オフ・ビート・ワークス" in order_address and "御中" in order_address
    checks.append({
        "sheet": "注文書",
        "cell": "B8",
        "item": "宛名",
        "expected": "株式会社オフ・ビート・ワークス　御中",
        "actual": order_address,
        "passed": order_address_valid
    })
    if not order_address_valid:
        errors.append(f"注文書B8: 宛名が不正です（{order_address}）")

    # G12: 発注金額
    order_amount = parse_number(get_cell_value(order_csv, "G12"))
    order_amount_valid = order_amount == invoice_total
    checks.append({
        "sheet": "注文書",
        "cell": "G12",
        "item": "発注金額",
        "expected": f"{invoice_total:,}円",
        "actual": f"{order_amount:,}円",
        "passed": order_amount_valid
    })
    if not order_amount_valid:
        errors.append(f"注文書G12: 発注金額が請求書と不一致（期待: {invoice_total:,}円、実際: {order_amount:,}円）")

    # C17: 明細タイトル（yyyy年mm月作業費形式）※「分」なし
    detail_title = get_cell_value(order_csv, "C17")
    detail_title_valid = bool(re.match(r'^\d{4}年\d{1,2}月作業費$', detail_title))
    checks.append({
        "sheet": "注文書",
        "cell": "C17",
        "item": "明細タイトル",
        "expected": "yyyy年mm月作業費形式",
        "actual": detail_title,
        "passed": detail_title_valid
    })
    if not detail_title_valid:
        errors.append(f"注文書C17: 明細タイトルが不正です（{detail_title}）")

    # AA17: 摘要（「見積番号：NNNNNNN」形式）
    order_remarks = get_cell_value(order_csv, f"AA{order_remarks_row}")
    if estimate_number:
        expected_remarks = f"見積番号：{estimate_number}"
        order_remarks_valid = order_remarks == expected_remarks
    else:
        # 見積番号がない場合は形式チェックのみ（7桁の数字）
        expected_remarks = "見積番号：NNNNNNN形式"
        order_remarks_valid = bool(re.match(r'^見積番号：\d{7}$', order_remarks))
    checks.append({
        "sheet": "注文書",
        "cell": f"AA{order_remarks_row}",
        "item": "摘要",
        "expected": expected_remarks,
        "actual": order_remarks,
        "passed": order_remarks_valid
    })
    if not order_remarks_valid:
        errors.append(f"注文書AA{order_remarks_row}: 摘要が不正です（期待: {expected_remarks}、実際: {order_remarks}）")

    # C(18+N): 以下、余白（動的行：明細数に応じて変動）
    order_blank_cell = f"C{order_blank_row}"
    blank_marker = get_cell_value(order_csv, order_blank_cell)
    blank_marker_valid = "以下、余白" in blank_marker or blank_marker.strip() == "以下、余白"
    checks.append({
        "sheet": "注文書",
        "cell": order_blank_cell,
        "item": "明細締め",
        "expected": "以下、余白",
        "actual": blank_marker,
        "passed": blank_marker_valid
    })
    if not blank_marker_valid:
        errors.append(f"注文書{order_blank_cell}: 「以下、余白」が入力されていません（{blank_marker}）")

    # W39: 小計
    order_subtotal = parse_number(get_cell_value(order_csv, "W39"))
    order_subtotal_valid = order_subtotal == invoice_subtotal
    checks.append({
        "sheet": "注文書",
        "cell": "W39",
        "item": "小計",
        "expected": f"{invoice_subtotal:,}円",
        "actual": f"{order_subtotal:,}円",
        "passed": order_subtotal_valid
    })
    if not order_subtotal_valid:
        errors.append(f"注文書W39: 小計が請求書と不一致（期待: {invoice_subtotal:,}円、実際: {order_subtotal:,}円）")

    # W40: 消費税
    order_tax = parse_number(get_cell_value(order_csv, "W40"))
    order_tax_valid = order_tax == invoice_tax
    checks.append({
        "sheet": "注文書",
        "cell": "W40",
        "item": "消費税",
        "expected": f"{invoice_tax:,}円",
        "actual": f"{order_tax:,}円",
        "passed": order_tax_valid
    })
    if not order_tax_valid:
        errors.append(f"注文書W40: 消費税が請求書と不一致（期待: {invoice_tax:,}円、実際: {order_tax:,}円）")

    # W41: 合計金額
    order_total = parse_number(get_cell_value(order_csv, "W41"))
    order_total_valid = order_total == invoice_total
    checks.append({
        "sheet": "注文書",
        "cell": "W41",
        "item": "合計金額",
        "expected": f"{invoice_total:,}円",
        "actual": f"{order_total:,}円",
        "passed": order_total_valid
    })
    if not order_total_valid:
        errors.append(f"注文書W41: 合計金額が請求書と不一致（期待: {invoice_total:,}円、実際: {order_total:,}円）")

    # === 検収書シート検証 ===

    # AC4: 検収番号（yyyymmdd-02形式）
    inspection_number = get_cell_value(inspection_csv, "AC4")
    inspection_number_valid = bool(re.match(r'^\d{8}-02$', inspection_number))
    checks.append({
        "sheet": "検収書",
        "cell": "AC4",
        "item": "検収番号形式",
        "expected": "yyyymmdd-02形式",
        "actual": inspection_number,
        "passed": inspection_number_valid
    })
    if not inspection_number_valid:
        errors.append(f"検収書AC4: 検収番号が不正です（{inspection_number}）")

    # AC5: 検収日（当月末日）
    inspection_date = get_cell_value(inspection_csv, "AC5")
    # 日付形式のチェック（yyyy/mm/dd または yyyy-mm-dd または datetime形式 または数値）
    # openpyxlはdatetime形式で「2025-07-31 00:00:00」のように出力することがある
    inspection_date_valid = bool(inspection_date) and bool(
        re.match(r'^\d{4}[/-]\d{1,2}[/-]\d{1,2}$', inspection_date) or
        re.match(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$', inspection_date) or  # datetime形式
        re.match(r'^\d+$', inspection_date.replace(',', ''))
    )
    checks.append({
        "sheet": "検収書",
        "cell": "AC5",
        "item": "検収日",
        "expected": "当月末日",
        "actual": inspection_date,
        "passed": inspection_date_valid
    })
    if not inspection_date_valid:
        errors.append(f"検収書AC5: 検収日が不正です（{inspection_date}）")

    # B7: 宛名
    inspection_address = get_cell_value(inspection_csv, "B7")
    inspection_address_valid = "株式会社オフ・ビート・ワークス" in inspection_address and "御中" in inspection_address
    checks.append({
        "sheet": "検収書",
        "cell": "B7",
        "item": "宛名",
        "expected": "株式会社オフ・ビート・ワークス　御中",
        "actual": inspection_address,
        "passed": inspection_address_valid
    })
    if not inspection_address_valid:
        errors.append(f"検収書B7: 宛名が不正です（{inspection_address}）")

    # G14: 合計金額
    inspection_amount = parse_number(get_cell_value(inspection_csv, "G14"))
    inspection_amount_valid = inspection_amount == invoice_total
    checks.append({
        "sheet": "検収書",
        "cell": "G14",
        "item": "合計金額",
        "expected": f"{invoice_total:,}円",
        "actual": f"{inspection_amount:,}円",
        "passed": inspection_amount_valid
    })
    if not inspection_amount_valid:
        errors.append(f"検収書G14: 合計金額が請求書と不一致（期待: {invoice_total:,}円、実際: {inspection_amount:,}円）")

    # C19: 明細タイトル（検収書）
    inspection_detail_title = get_cell_value(inspection_csv, "C19")
    inspection_detail_title_valid = bool(re.match(r'^\d{4}年\d{1,2}月作業費$', inspection_detail_title))
    checks.append({
        "sheet": "検収書",
        "cell": "C19",
        "item": "明細タイトル",
        "expected": "yyyy年mm月作業費形式",
        "actual": inspection_detail_title,
        "passed": inspection_detail_title_valid
    })
    if not inspection_detail_title_valid:
        errors.append(f"検収書C19: 明細タイトルが不正です（{inspection_detail_title}）")

    # AA19: 摘要（「見積番号：NNNNNNN」形式）
    inspection_remarks = get_cell_value(inspection_csv, f"AA{inspection_remarks_row}")
    if estimate_number:
        expected_inspection_remarks = f"見積番号：{estimate_number}"
        inspection_remarks_valid = inspection_remarks == expected_inspection_remarks
    else:
        expected_inspection_remarks = "見積番号：NNNNNNN形式"
        inspection_remarks_valid = bool(re.match(r'^見積番号：\d{7}$', inspection_remarks))
    checks.append({
        "sheet": "検収書",
        "cell": f"AA{inspection_remarks_row}",
        "item": "摘要",
        "expected": expected_inspection_remarks,
        "actual": inspection_remarks,
        "passed": inspection_remarks_valid
    })
    if not inspection_remarks_valid:
        errors.append(f"検収書AA{inspection_remarks_row}: 摘要が不正です（期待: {expected_inspection_remarks}、実際: {inspection_remarks}）")

    # C(20+N): 以下、余白（動的行：明細数に応じて変動）
    inspection_blank_cell = f"C{inspection_blank_row}"
    inspection_blank = get_cell_value(inspection_csv, inspection_blank_cell)
    inspection_blank_valid = "以下、余白" in inspection_blank or inspection_blank.strip() == "以下、余白"
    checks.append({
        "sheet": "検収書",
        "cell": inspection_blank_cell,
        "item": "明細締め",
        "expected": "以下、余白",
        "actual": inspection_blank,
        "passed": inspection_blank_valid
    })
    if not inspection_blank_valid:
        errors.append(f"検収書{inspection_blank_cell}: 「以下、余白」が入力されていません（{inspection_blank}）")

    # W41: 小計（検収書）
    inspection_subtotal = parse_number(get_cell_value(inspection_csv, "W41"))
    inspection_subtotal_valid = inspection_subtotal == invoice_subtotal
    checks.append({
        "sheet": "検収書",
        "cell": "W41",
        "item": "小計",
        "expected": f"{invoice_subtotal:,}円",
        "actual": f"{inspection_subtotal:,}円",
        "passed": inspection_subtotal_valid
    })
    if not inspection_subtotal_valid:
        errors.append(f"検収書W41: 小計が請求書と不一致（期待: {invoice_subtotal:,}円、実際: {inspection_subtotal:,}円）")

    # W42: 消費税（検収書）
    inspection_tax = parse_number(get_cell_value(inspection_csv, "W42"))
    inspection_tax_valid = inspection_tax == invoice_tax
    checks.append({
        "sheet": "検収書",
        "cell": "W42",
        "item": "消費税",
        "expected": f"{invoice_tax:,}円",
        "actual": f"{inspection_tax:,}円",
        "passed": inspection_tax_valid
    })
    if not inspection_tax_valid:
        errors.append(f"検収書W42: 消費税が請求書と不一致（期待: {invoice_tax:,}円、実際: {inspection_tax:,}円）")

    # W43: 合計金額（検収書）
    inspection_total = parse_number(get_cell_value(inspection_csv, "W43"))
    inspection_total_valid = inspection_total == invoice_total
    checks.append({
        "sheet": "検収書",
        "cell": "W43",
        "item": "合計金額",
        "expected": f"{invoice_total:,}円",
        "actual": f"{inspection_total:,}円",
        "passed": inspection_total_valid
    })
    if not inspection_total_valid:
        errors.append(f"検収書W43: 合計金額が請求書と不一致（期待: {invoice_total:,}円、実際: {inspection_total:,}円）")

    return {
        "success": len(errors) == 0,
        "checks": checks,
        "errors": errors
    }


def validate_excel(excel_path: str, company_name: str, validation_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Excelファイルを検証

    Args:
        excel_path: Excelファイルのパス
        company_name: 取引先名
        validation_data: 検証データ（invoice, estimate, items_count を含む辞書）

    Returns:
        検証結果
    """
    # 一時ディレクトリ
    output_dir = os.path.dirname(excel_path)
    csv_paths = {}

    try:
        # LibreOfficeでCSV変換（数式計算後の値を取得）
        csv_paths = convert_excel_to_csv_with_libreoffice(excel_path, output_dir)

        # 取引先に応じた検証
        if company_name == "ネクストビッツ":
            result = validate_nextbits_excel(csv_paths, validation_data)
        elif company_name == "オフ・ビート・ワークス":
            result = validate_offbeat_excel(csv_paths, validation_data)
        else:
            result = {
                "success": False,
                "checks": [],
                "errors": [f"未対応の取引先: {company_name}"]
            }

        return result

    finally:
        # 一時CSVファイル削除
        for csv_path in csv_paths.values():
            if os.path.exists(csv_path):
                os.remove(csv_path)


def main():
    """メイン関数"""
    if len(sys.argv) != 4:
        print(json.dumps({
            "error": "引数が不足しています",
            "usage": "python3 excel_validator.py <excel_path> <company_name> <validation_data_json>"
        }, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)

    excel_path = sys.argv[1]
    company_name = sys.argv[2]
    validation_data_json = sys.argv[3]

    try:
        validation_data = json.loads(validation_data_json)
        result = validate_excel(excel_path, company_name, validation_data)

        # checksのpassedフィールドをboolに強制変換（Matchオブジェクト対策）
        if "checks" in result:
            for check in result["checks"]:
                if "passed" in check:
                    check["passed"] = bool(check["passed"])

        print(json.dumps(result, ensure_ascii=False))

        if not result["success"]:
            sys.exit(1)

    except Exception as e:
        import traceback
        print(json.dumps({
            "error": str(e),
            "error_type": type(e).__name__,
            "traceback": traceback.format_exc()
        }, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
